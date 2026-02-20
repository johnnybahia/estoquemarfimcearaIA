#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Relatórios Inteligentes - Marfim Estoque Ceará
Autor: Johnny
Data: 2026-02-13

Relatórios completos com:
- 📊 Curva ABC (classificação por valor/movimento)
- 📈 Análise de giro de estoque
- 📉 Itens parados / sem movimento
- 💰 Resumo financeiro por grupo
- 📅 Relatório diário/semanal/mensal
- 📄 Export Excel (openpyxl)
- 📋 Export PDF (fpdf2)
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass, field
from enum import Enum
from collections import defaultdict

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ========================================
# ENUMS E CLASSES
# ========================================

class ClasseABC(Enum):
    """Classe Curva ABC"""
    A = "A"   # 80% do volume → top itens
    B = "B"   # 15% do volume → itens médios
    C = "C"   # 5%  do volume → itens pouco movimentados


class PeriodoRelatorio(Enum):
    """Períodos disponíveis"""
    HOJE = "hoje"
    SEMANA = "semana"
    MES = "mes"
    TRIMESTRE = "trimestre"
    PERSONALIZADO = "personalizado"


@dataclass
class ItemCurvaABC:
    """Item classificado na Curva ABC"""
    item: str
    grupo: str
    saldo: float
    total_movimentacoes: int
    total_entradas: float
    total_saidas: float
    volume_total: float      # entradas + saidas
    percentual_volume: float
    percentual_acumulado: float
    classe: ClasseABC

    def to_dict(self) -> Dict[str, Any]:
        return {
            'item': self.item,
            'grupo': self.grupo,
            'saldo': self.saldo,
            'total_movimentacoes': self.total_movimentacoes,
            'total_entradas': self.total_entradas,
            'total_saidas': self.total_saidas,
            'volume_total': self.volume_total,
            'percentual_volume': round(self.percentual_volume, 2),
            'percentual_acumulado': round(self.percentual_acumulado, 2),
            'classe': self.classe.value
        }


@dataclass
class ResumoGrupo:
    """Resumo de um grupo de itens"""
    grupo: str
    total_itens: int
    saldo_total: float
    total_movimentacoes: int
    total_entradas: float
    total_saidas: float
    itens_zerados: int
    itens_criticos: int
    classe_predominante: str

    def to_dict(self) -> Dict[str, Any]:
        return {
            'grupo': self.grupo,
            'total_itens': self.total_itens,
            'saldo_total': round(self.saldo_total, 2),
            'total_movimentacoes': self.total_movimentacoes,
            'total_entradas': round(self.total_entradas, 2),
            'total_saidas': round(self.total_saidas, 2),
            'itens_zerados': self.itens_zerados,
            'itens_criticos': self.itens_criticos,
            'classe_predominante': self.classe_predominante
        }


@dataclass
class RelatorioCompleto:
    """Relatório completo do estoque"""
    periodo: str
    data_geracao: str
    total_itens: int
    total_grupos: int
    saldo_geral: float
    total_movimentacoes: int
    total_entradas: float
    total_saidas: float
    itens_zerados: int
    itens_sem_movimento: int
    curva_abc: Dict[str, List]  # {A: [...], B: [...], C: [...]}
    resumo_grupos: List[ResumoGrupo]
    top_movimentados: List[Dict]
    itens_parados: List[Dict]
    grafico_abc: Dict[str, Any]     # dados para Chart.js
    grafico_grupos: Dict[str, Any]  # dados para Chart.js
    grafico_movimentacoes: Dict[str, Any]

    def to_dict(self) -> Dict[str, Any]:
        return {
            'periodo': self.periodo,
            'data_geracao': self.data_geracao,
            'total_itens': self.total_itens,
            'total_grupos': self.total_grupos,
            'saldo_geral': round(self.saldo_geral, 2),
            'total_movimentacoes': self.total_movimentacoes,
            'total_entradas': round(self.total_entradas, 2),
            'total_saidas': round(self.total_saidas, 2),
            'itens_zerados': self.itens_zerados,
            'itens_sem_movimento': self.itens_sem_movimento,
            'curva_abc': {
                k: [i.to_dict() for i in v]
                for k, v in self.curva_abc.items()
            },
            'resumo_grupos': [g.to_dict() for g in self.resumo_grupos],
            'top_movimentados': self.top_movimentados,
            'itens_parados': self.itens_parados,
            'grafico_abc': self.grafico_abc,
            'grafico_grupos': self.grafico_grupos,
            'grafico_movimentacoes': self.grafico_movimentacoes
        }


# ========================================
# CONFIGURAÇÕES
# ========================================

class ConfigRelatorios:
    """Configurações de relatórios"""
    # Curva ABC
    LIMITE_CLASSE_A = 80.0   # % acumulado para classe A
    LIMITE_CLASSE_B = 95.0   # % acumulado para classe B (B = 95-80)
    # C = tudo acima de 95%

    # Parado
    DIAS_SEM_MOVIMENTO_PARADO = 30  # Dias sem movimento = parado

    # Top
    TOP_N_MOVIMENTADOS = 10


# ========================================
# CLASSE PRINCIPAL
# ========================================

class GerenciadorRelatorios:
    """
    Gerenciador de relatórios inteligentes

    Uso:
        gerenciador = GerenciadorRelatorios()
        relatorio = gerenciador.gerar_relatorio_completo()
        curva_abc = gerenciador.calcular_curva_abc()
    """

    def __init__(self):
        logger.info("✅ GerenciadorRelatorios inicializado")

    def _obter_dados_indice(self) -> List[Dict]:
        """Obtém todos os itens do índice"""
        try:
            from indice_otimizado import indice_otimizado
            itens = indice_otimizado.obter_todos_itens()
            return itens if itens else []
        except Exception as e:
            logger.warning(f"Índice indisponível: {e}")
            return []

    def _obter_historico_periodo(
        self,
        data_inicio: Optional[str] = None,
        data_fim: Optional[str] = None,
        limite: int = 500
    ) -> List[Dict]:
        """Obtém histórico de movimentações"""
        try:
            from historico_otimizado import gerenciador_historico

            if data_inicio and data_fim:
                registros = gerenciador_historico.buscar_por_periodo(
                    data_inicio, data_fim, limite
                )
            else:
                registros = gerenciador_historico.obter_ultimos(limite)

            return [r.to_dict() for r in registros]
        except Exception as e:
            logger.warning(f"Histórico indisponível: {e}")
            return []

    def calcular_curva_abc(
        self,
        historico: Optional[List[Dict]] = None
    ) -> Dict[str, List[ItemCurvaABC]]:
        """
        Calcula Curva ABC baseada em volume de movimentações

        Classificação:
        - A: itens que representam 80% do volume → prioridade máxima
        - B: itens que representam 15% adicionais → prioridade média
        - C: restantes 5% → prioridade baixa

        Args:
            historico: Lista de movimentações (se None, busca)

        Returns:
            Dict com {'A': [...], 'B': [...], 'C': [...]}
        """
        if historico is None:
            historico = self._obter_historico_periodo(limite=500)

        # Agrega volume por item
        volume_por_item: Dict[str, Dict] = defaultdict(lambda: {
            'entradas': 0.0,
            'saidas': 0.0,
            'movimentacoes': 0,
            'grupo': ''
        })

        for reg in historico:
            item = reg.get('item', '')
            qtd = abs(float(reg.get('quantidade', 0)))
            tipo = reg.get('tipo_movimentacao', '').upper()
            grupo = reg.get('grupo', '')

            volume_por_item[item]['movimentacoes'] += 1
            volume_por_item[item]['grupo'] = grupo

            if 'ENTRADA' in tipo:
                volume_por_item[item]['entradas'] += qtd
            else:
                volume_por_item[item]['saidas'] += qtd

        if not volume_por_item:
            return {'A': [], 'B': [], 'C': []}

        # Busca saldos do índice
        saldos = {}
        try:
            from indice_otimizado import indice_otimizado
            for item_nome in volume_por_item:
                dados = indice_otimizado.buscar_item(item_nome)
                if dados:
                    saldos[item_nome] = float(dados.get('saldo', 0))
        except:
            pass

        # Calcula volume total por item (entradas + saídas)
        volumes = []
        for item_nome, dados in volume_por_item.items():
            vol = dados['entradas'] + dados['saidas']
            volumes.append((item_nome, vol, dados))

        # Ordena por volume (maior primeiro)
        volumes.sort(key=lambda x: x[1], reverse=True)

        # Calcula total geral
        total_volume = sum(v[1] for v in volumes)

        if total_volume == 0:
            return {'A': [], 'B': [], 'C': []}

        # Classifica ABC
        resultado = {'A': [], 'B': [], 'C': []}
        acumulado = 0.0

        for item_nome, vol, dados in volumes:
            percentual = (vol / total_volume) * 100
            acumulado += percentual

            if acumulado <= ConfigRelatorios.LIMITE_CLASSE_A:
                classe = ClasseABC.A
            elif acumulado <= ConfigRelatorios.LIMITE_CLASSE_B:
                classe = ClasseABC.B
            else:
                classe = ClasseABC.C

            item_abc = ItemCurvaABC(
                item=item_nome,
                grupo=dados['grupo'],
                saldo=saldos.get(item_nome, 0.0),
                total_movimentacoes=dados['movimentacoes'],
                total_entradas=dados['entradas'],
                total_saidas=dados['saidas'],
                volume_total=vol,
                percentual_volume=percentual,
                percentual_acumulado=acumulado,
                classe=classe
            )

            resultado[classe.value].append(item_abc)

        return resultado

    def calcular_resumo_grupos(
        self,
        historico: Optional[List[Dict]] = None
    ) -> List[ResumoGrupo]:
        """
        Calcula resumo por grupo de itens

        Returns:
            Lista de ResumoGrupo ordenada por volume
        """
        if historico is None:
            historico = self._obter_historico_periodo(limite=500)

        grupos: Dict[str, Dict] = defaultdict(lambda: {
            'itens': set(),
            'movimentacoes': 0,
            'entradas': 0.0,
            'saidas': 0.0
        })

        for reg in historico:
            grupo = reg.get('grupo', 'SEM_GRUPO') or 'SEM_GRUPO'
            item = reg.get('item', '')
            qtd = abs(float(reg.get('quantidade', 0)))
            tipo = reg.get('tipo_movimentacao', '').upper()

            grupos[grupo]['itens'].add(item)
            grupos[grupo]['movimentacoes'] += 1

            if 'ENTRADA' in tipo:
                grupos[grupo]['entradas'] += qtd
            else:
                grupos[grupo]['saidas'] += qtd

        # Calcula curva ABC para classificação
        curva = self.calcular_curva_abc(historico)

        # Mapa item → classe
        mapa_classe = {}
        for classe, itens in curva.items():
            for item_abc in itens:
                mapa_classe[item_abc.item] = classe

        # Busca saldos
        saldos_por_grupo: Dict[str, float] = defaultdict(float)
        zerados_por_grupo: Dict[str, int] = defaultdict(int)
        criticos_por_grupo: Dict[str, int] = defaultdict(int)

        try:
            from indice_otimizado import indice_otimizado
            for item_nome in set(i for g in grupos.values() for i in g['itens']):
                dados = indice_otimizado.buscar_item(item_nome)
                if dados:
                    saldo = float(dados.get('saldo', 0))
                    grupo = dados.get('grupo', 'SEM_GRUPO') or 'SEM_GRUPO'
                    saldos_por_grupo[grupo] += saldo
                    if saldo <= 0:
                        zerados_por_grupo[grupo] += 1
                    if saldo < 10:
                        criticos_por_grupo[grupo] += 1
        except:
            pass

        # Monta resumo
        resumos = []
        for grupo_nome, dados in grupos.items():
            # Classe predominante
            classes_grupo = [mapa_classe.get(i, 'C') for i in dados['itens']]
            classe_pred = max(set(classes_grupo), key=classes_grupo.count) if classes_grupo else 'C'

            resumos.append(ResumoGrupo(
                grupo=grupo_nome,
                total_itens=len(dados['itens']),
                saldo_total=saldos_por_grupo.get(grupo_nome, 0.0),
                total_movimentacoes=dados['movimentacoes'],
                total_entradas=dados['entradas'],
                total_saidas=dados['saidas'],
                itens_zerados=zerados_por_grupo.get(grupo_nome, 0),
                itens_criticos=criticos_por_grupo.get(grupo_nome, 0),
                classe_predominante=classe_pred
            ))

        # Ordena por volume total
        resumos.sort(key=lambda r: r.total_entradas + r.total_saidas, reverse=True)
        return resumos

    def obter_itens_parados(self, dias: int = None) -> List[Dict]:
        """
        Retorna itens sem movimentação no período

        Args:
            dias: Dias sem movimento para considerar parado

        Returns:
            Lista de itens parados
        """
        dias = dias or ConfigRelatorios.DIAS_SEM_MOVIMENTO_PARADO

        try:
            from alertas_automaticos import gerenciador_alertas
            alertas = gerenciador_alertas.obter_todos_alertas()

            parados = []
            for alerta in alertas:
                if alerta.get('status') in ('critico', 'atencao'):
                    dias_sem_mov = alerta.get('dias_sem_movimento', 0)
                    if isinstance(dias_sem_mov, (int, float)) and dias_sem_mov >= dias:
                        parados.append({
                            'item': alerta.get('item', ''),
                            'grupo': alerta.get('grupo', ''),
                            'saldo': alerta.get('saldo', 0),
                            'dias_sem_movimento': dias_sem_mov,
                            'ultima_movimentacao': alerta.get('ultima_movimentacao', 'N/A'),
                            'status': alerta.get('status', '')
                        })

            parados.sort(key=lambda x: x.get('dias_sem_movimento', 0), reverse=True)
            return parados

        except Exception as e:
            logger.warning(f"Erro ao obter parados: {e}")
            return []

    def _gerar_dados_grafico_abc(
        self,
        curva: Dict[str, List[ItemCurvaABC]]
    ) -> Dict[str, Any]:
        """Gera dados para gráfico de pizza ABC (Chart.js)"""
        return {
            'type': 'doughnut',
            'labels': ['Classe A (80%)', 'Classe B (15%)', 'Classe C (5%)'],
            'data': [
                len(curva.get('A', [])),
                len(curva.get('B', [])),
                len(curva.get('C', []))
            ],
            'colors': ['#4CAF50', '#FF9800', '#F44336'],
            'legenda': {
                'A': f"{len(curva.get('A', []))} itens (alta prioridade)",
                'B': f"{len(curva.get('B', []))} itens (média prioridade)",
                'C': f"{len(curva.get('C', []))} itens (baixa prioridade)"
            }
        }

    def _gerar_dados_grafico_grupos(
        self,
        resumos: List[ResumoGrupo]
    ) -> Dict[str, Any]:
        """Gera dados para gráfico de barras por grupo (Chart.js)"""
        top = resumos[:8]
        return {
            'type': 'bar',
            'labels': [r.grupo for r in top],
            'datasets': [
                {
                    'label': 'Entradas',
                    'data': [r.total_entradas for r in top],
                    'color': '#4CAF50'
                },
                {
                    'label': 'Saídas',
                    'data': [r.total_saidas for r in top],
                    'color': '#F44336'
                }
            ]
        }

    def _gerar_dados_grafico_movimentacoes(
        self,
        historico: List[Dict]
    ) -> Dict[str, Any]:
        """Gera dados para gráfico de linha (movimentações por dia)"""
        por_dia: Dict[str, Dict] = defaultdict(lambda: {'entradas': 0.0, 'saidas': 0.0})

        for reg in historico:
            data = reg.get('data', '')
            qtd = abs(float(reg.get('quantidade', 0)))
            tipo = reg.get('tipo_movimentacao', '').upper()

            if 'ENTRADA' in tipo:
                por_dia[data]['entradas'] += qtd
            else:
                por_dia[data]['saidas'] += qtd

        # Ordena por data
        datas_ord = sorted(por_dia.keys())[-14:]  # Últimos 14 dias

        return {
            'type': 'line',
            'labels': datas_ord,
            'datasets': [
                {
                    'label': 'Entradas',
                    'data': [por_dia[d]['entradas'] for d in datas_ord],
                    'color': '#4CAF50'
                },
                {
                    'label': 'Saídas',
                    'data': [por_dia[d]['saidas'] for d in datas_ord],
                    'color': '#F44336'
                }
            ]
        }

    def gerar_relatorio_completo(
        self,
        periodo: str = 'mes',
        data_inicio: Optional[str] = None,
        data_fim: Optional[str] = None
    ) -> RelatorioCompleto:
        """
        Gera relatório completo do estoque

        Args:
            periodo: hoje/semana/mes/trimestre/personalizado
            data_inicio: Se personalizado, data início (DD/MM/YYYY)
            data_fim: Se personalizado, data fim (DD/MM/YYYY)

        Returns:
            RelatorioCompleto
        """
        # Define período
        hoje = datetime.now()
        if periodo == 'hoje':
            data_inicio = hoje.strftime('%d/%m/%Y')
            data_fim = data_inicio
        elif periodo == 'semana':
            data_inicio = (hoje - timedelta(days=7)).strftime('%d/%m/%Y')
            data_fim = hoje.strftime('%d/%m/%Y')
        elif periodo == 'mes':
            data_inicio = (hoje - timedelta(days=30)).strftime('%d/%m/%Y')
            data_fim = hoje.strftime('%d/%m/%Y')
        elif periodo == 'trimestre':
            data_inicio = (hoje - timedelta(days=90)).strftime('%d/%m/%Y')
            data_fim = hoje.strftime('%d/%m/%Y')

        # Busca histórico do período
        historico = self._obter_historico_periodo(data_inicio, data_fim, limite=1000)

        # Calcula Curva ABC
        curva = self.calcular_curva_abc(historico)

        # Resumo por grupos
        resumos_grupos = self.calcular_resumo_grupos(historico)

        # Itens parados
        itens_parados = self.obter_itens_parados()

        # Estatísticas gerais
        total_entradas = sum(
            abs(float(r.get('quantidade', 0)))
            for r in historico
            if 'ENTRADA' in r.get('tipo_movimentacao', '').upper()
        )
        total_saidas = sum(
            abs(float(r.get('quantidade', 0)))
            for r in historico
            if 'SAIDA' in r.get('tipo_movimentacao', '').upper()
        )

        # Top movimentados
        todos_itens_abc = (
            curva.get('A', []) + curva.get('B', []) + curva.get('C', [])
        )
        top_movimentados = [
            i.to_dict() for i in
            sorted(todos_itens_abc, key=lambda x: x.volume_total, reverse=True)
            [:ConfigRelatorios.TOP_N_MOVIMENTADOS]
        ]

        # Saldo geral
        saldo_geral = sum(i.saldo for i in todos_itens_abc)
        itens_zerados = sum(1 for i in todos_itens_abc if i.saldo <= 0)

        # Gráficos
        grafico_abc = self._gerar_dados_grafico_abc(curva)
        grafico_grupos = self._gerar_dados_grafico_grupos(resumos_grupos)
        grafico_mov = self._gerar_dados_grafico_movimentacoes(historico)

        return RelatorioCompleto(
            periodo=periodo,
            data_geracao=hoje.strftime('%d/%m/%Y %H:%M'),
            total_itens=len(todos_itens_abc),
            total_grupos=len(resumos_grupos),
            saldo_geral=saldo_geral,
            total_movimentacoes=len(historico),
            total_entradas=total_entradas,
            total_saidas=total_saidas,
            itens_zerados=itens_zerados,
            itens_sem_movimento=len(itens_parados),
            curva_abc=curva,
            resumo_grupos=resumos_grupos,
            top_movimentados=top_movimentados,
            itens_parados=itens_parados,
            grafico_abc=grafico_abc,
            grafico_grupos=grafico_grupos,
            grafico_movimentacoes=grafico_mov
        )

    # ========================================
    # EXPORTS
    # ========================================

    def exportar_excel(self, relatorio: RelatorioCompleto) -> bytes:
        """
        Exporta relatório para Excel (.xlsx)

        Returns:
            Bytes do arquivo Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io

            wb = openpyxl.Workbook()

            # ========== ABA 1: Resumo ==========
            ws = wb.active
            ws.title = "📊 Resumo"
            self._excel_aba_resumo(ws, relatorio)

            # ========== ABA 2: Curva ABC ==========
            ws_abc = wb.create_sheet("🔤 Curva ABC")
            self._excel_aba_abc(ws_abc, relatorio)

            # ========== ABA 3: Por Grupos ==========
            ws_grupos = wb.create_sheet("📦 Por Grupos")
            self._excel_aba_grupos(ws_grupos, relatorio)

            # ========== ABA 4: Itens Parados ==========
            ws_parados = wb.create_sheet("⏸️ Itens Parados")
            self._excel_aba_parados(ws_parados, relatorio)

            # ========== ABA 5: Top Movimentados ==========
            ws_top = wb.create_sheet("🏆 Top Movimentados")
            self._excel_aba_top(ws_top, relatorio)

            # Salva em bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.read()

        except ImportError:
            logger.error("openpyxl não instalado. Execute: pip install openpyxl")
            raise
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {e}")
            raise

    def _excel_estilo_header(self, ws, row, cols, texto_lista, cor_fundo='4472C4'):
        """Aplica estilo de cabeçalho"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            fill = PatternFill("solid", fgColor=cor_fundo)
            font = Font(bold=True, color="FFFFFF")
            align = Alignment(horizontal="center", vertical="center")

            for col, texto in zip(cols, texto_lista):
                cell = ws.cell(row=row, column=col, value=texto)
                cell.fill = fill
                cell.font = font
                cell.alignment = align
        except Exception as e:
            logger.debug(f"Erro ao aplicar estilo: {e}")

    def _excel_aba_resumo(self, ws, relatorio: RelatorioCompleto):
        """Preenche aba de resumo"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20

            # Título
            ws['A1'] = '📊 RELATÓRIO DE ESTOQUE - MARFIM'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f'Gerado em: {relatorio.data_geracao}'
            ws['A3'] = f'Período: {relatorio.periodo.upper()}'

            ws['A5'] = 'INDICADOR'
            ws['B5'] = 'VALOR'
            self._excel_estilo_header(ws, 5, [1, 2], ['INDICADOR', 'VALOR'])

            dados = [
                ('Total de Itens', relatorio.total_itens),
                ('Total de Grupos', relatorio.total_grupos),
                ('Saldo Geral', f'{relatorio.saldo_geral:.2f}'),
                ('Total Movimentações', relatorio.total_movimentacoes),
                ('Total Entradas', f'{relatorio.total_entradas:.2f}'),
                ('Total Saídas', f'{relatorio.total_saidas:.2f}'),
                ('Itens Zerados', relatorio.itens_zerados),
                ('Itens Parados (30+ dias)', relatorio.itens_sem_movimento),
                ('Itens Classe A', len(relatorio.curva_abc.get('A', []))),
                ('Itens Classe B', len(relatorio.curva_abc.get('B', []))),
                ('Itens Classe C', len(relatorio.curva_abc.get('C', []))),
            ]

            for i, (indicador, valor) in enumerate(dados, start=6):
                ws.cell(row=i, column=1, value=indicador)
                ws.cell(row=i, column=2, value=valor)
        except Exception as e:
            logger.debug(f"Erro resumo excel: {e}")

    def _excel_aba_abc(self, ws, relatorio: RelatorioCompleto):
        """Preenche aba Curva ABC"""
        try:
            from openpyxl.styles import PatternFill

            headers = ['Item', 'Grupo', 'Classe', 'Volume Total',
                       'Entradas', 'Saídas', 'Saldo', '% Volume', '% Acumulado']
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 18

            self._excel_estilo_header(ws, 1, range(1, 10), headers)

            row = 2
            cores = {'A': 'E8F5E9', 'B': 'FFF3E0', 'C': 'FFEBEE'}
            for classe in ['A', 'B', 'C']:
                for item in relatorio.curva_abc.get(classe, []):
                    d = item.to_dict()
                    fill = PatternFill("solid", fgColor=cores[classe])
                    vals = [
                        d['item'], d['grupo'], d['classe'],
                        d['volume_total'], d['total_entradas'], d['total_saidas'],
                        d['saldo'], d['percentual_volume'], d['percentual_acumulado']
                    ]
                    for col, val in enumerate(vals, start=1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.fill = fill
                    row += 1
        except Exception as e:
            logger.debug(f"Erro abc excel: {e}")

    def _excel_aba_grupos(self, ws, relatorio: RelatorioCompleto):
        """Preenche aba de grupos"""
        headers = ['Grupo', 'Total Itens', 'Saldo Total', 'Movimentações',
                   'Entradas', 'Saídas', 'Itens Zerados', 'Itens Críticos', 'Classe']
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 18

        self._excel_estilo_header(ws, 1, range(1, 10), headers, cor_fundo='2E7D32')

        for row, grupo in enumerate(relatorio.resumo_grupos, start=2):
            d = grupo.to_dict()
            vals = [
                d['grupo'], d['total_itens'], d['saldo_total'],
                d['total_movimentacoes'], d['total_entradas'], d['total_saidas'],
                d['itens_zerados'], d['itens_criticos'], d['classe_predominante']
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=row, column=col, value=val)

    def _excel_aba_parados(self, ws, relatorio: RelatorioCompleto):
        """Preenche aba de itens parados"""
        headers = ['Item', 'Grupo', 'Saldo', 'Dias Sem Movimento',
                   'Última Movimentação', 'Status']
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 22

        self._excel_estilo_header(ws, 1, range(1, 7), headers, cor_fundo='C62828')

        for row, item in enumerate(relatorio.itens_parados, start=2):
            vals = [
                item.get('item', ''), item.get('grupo', ''), item.get('saldo', 0),
                item.get('dias_sem_movimento', 0), item.get('ultima_movimentacao', ''),
                item.get('status', '')
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=row, column=col, value=val)

    def _excel_aba_top(self, ws, relatorio: RelatorioCompleto):
        """Preenche aba de top movimentados"""
        headers = ['#', 'Item', 'Grupo', 'Volume Total', 'Entradas', 'Saídas',
                   'Movimentações', 'Saldo', 'Classe']
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 18

        self._excel_estilo_header(ws, 1, range(1, 10), headers, cor_fundo='1565C0')

        for row, item in enumerate(relatorio.top_movimentados, start=2):
            vals = [
                row - 1, item.get('item', ''), item.get('grupo', ''),
                item.get('volume_total', 0), item.get('total_entradas', 0),
                item.get('total_saidas', 0), item.get('total_movimentacoes', 0),
                item.get('saldo', 0), item.get('classe', '')
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=row, column=col, value=val)

    def exportar_pdf(self, relatorio: RelatorioCompleto) -> bytes:
        """
        Exporta relatório para PDF

        Returns:
            Bytes do arquivo PDF
        """
        try:
            from fpdf import FPDF
            import io

            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()

            # Título
            pdf.set_font('Helvetica', 'B', 20)
            pdf.set_fill_color(102, 126, 234)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 15, 'RELATORIO DE ESTOQUE - MARFIM', fill=True, ln=True, align='C')

            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 8, f'Gerado em: {relatorio.data_geracao}  |  Periodo: {relatorio.periodo.upper()}', ln=True, align='C')
            pdf.ln(5)

            # Resumo Geral
            self._pdf_secao(pdf, 'RESUMO GERAL')
            dados_resumo = [
                ('Total de Itens', str(relatorio.total_itens)),
                ('Total de Grupos', str(relatorio.total_grupos)),
                ('Saldo Geral', f'{relatorio.saldo_geral:.2f}'),
                ('Total Movimentacoes', str(relatorio.total_movimentacoes)),
                ('Total Entradas', f'{relatorio.total_entradas:.2f}'),
                ('Total Saidas', f'{relatorio.total_saidas:.2f}'),
                ('Itens Zerados', str(relatorio.itens_zerados)),
                ('Itens Parados', str(relatorio.itens_sem_movimento)),
            ]
            self._pdf_tabela_kv(pdf, dados_resumo)
            pdf.ln(5)

            # Curva ABC
            self._pdf_secao(pdf, 'CURVA ABC')
            abc_dados = [
                ('Classe A (alta prioridade)', str(len(relatorio.curva_abc.get('A', [])))),
                ('Classe B (media prioridade)', str(len(relatorio.curva_abc.get('B', [])))),
                ('Classe C (baixa prioridade)', str(len(relatorio.curva_abc.get('C', [])))),
            ]
            self._pdf_tabela_kv(pdf, abc_dados)
            pdf.ln(5)

            # Top 10 Movimentados
            self._pdf_secao(pdf, 'TOP 10 MAIS MOVIMENTADOS')
            headers = ['Item', 'Classe', 'Volume', 'Saldo']
            rows = [
                [
                    i.get('item', '')[:25],
                    i.get('classe', ''),
                    f"{i.get('volume_total', 0):.1f}",
                    f"{i.get('saldo', 0):.1f}"
                ]
                for i in relatorio.top_movimentados
            ]
            self._pdf_tabela(pdf, headers, rows)
            pdf.ln(5)

            # Itens Parados
            if relatorio.itens_parados:
                self._pdf_secao(pdf, 'ITENS PARADOS (30+ DIAS)')
                headers_p = ['Item', 'Grupo', 'Dias Parado', 'Saldo']
                rows_p = [
                    [
                        p.get('item', '')[:25],
                        p.get('grupo', '')[:15],
                        str(p.get('dias_sem_movimento', 0)),
                        f"{p.get('saldo', 0):.1f}"
                    ]
                    for p in relatorio.itens_parados[:15]
                ]
                self._pdf_tabela(pdf, headers_p, rows_p)

            return bytes(pdf.output())

        except ImportError:
            logger.error("fpdf2 não instalado. Execute: pip install fpdf2")
            raise
        except Exception as e:
            logger.error(f"Erro ao gerar PDF: {e}")
            raise

    def _pdf_secao(self, pdf, titulo: str):
        """Adiciona seção ao PDF"""
        pdf.set_fill_color(102, 126, 234)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, titulo, fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 10)

    def _pdf_tabela_kv(self, pdf, dados: List[Tuple]):
        """Tabela chave-valor simples"""
        for chave, valor in dados:
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(80, 6, chave + ':', border='B')
            pdf.set_font('Helvetica', '', 9)
            pdf.cell(0, 6, valor, border='B', ln=True)

    def _pdf_tabela(self, pdf, headers: List[str], rows: List[List]):
        """Tabela com cabeçalho"""
        col_w = 190 // len(headers)

        # Cabeçalho
        pdf.set_fill_color(200, 200, 200)
        pdf.set_font('Helvetica', 'B', 9)
        for h in headers:
            pdf.cell(col_w, 6, h, border=1, fill=True)
        pdf.ln()

        # Linhas
        pdf.set_font('Helvetica', '', 8)
        for i, row in enumerate(rows):
            if i % 2 == 0:
                pdf.set_fill_color(245, 245, 245)
            else:
                pdf.set_fill_color(255, 255, 255)
            for cell in row:
                pdf.cell(col_w, 6, str(cell), border=1, fill=True)
            pdf.ln()


# ========================================
# SINGLETON GLOBAL
# ========================================
gerenciador_relatorios = GerenciadorRelatorios()


if __name__ == '__main__':
    print("🧪 Testando GerenciadorRelatorios...")
    relatorio = gerenciador_relatorios.gerar_relatorio_completo('mes')
    print(f"✅ Relatório gerado: {relatorio.total_itens} itens, {relatorio.total_movimentacoes} movimentações")
    print(f"   Curva A: {len(relatorio.curva_abc.get('A', []))} itens")
    print(f"   Curva B: {len(relatorio.curva_abc.get('B', []))} itens")
    print(f"   Curva C: {len(relatorio.curva_abc.get('C', []))} itens")
